import sqlite3

import pwinput
from time import sleep
from fpdf import FPDF
from openpyxl import Workbook
from tkinter import filedialog, messagebox

import sqlite3

def iniciaDB():
    try:
        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()

            # Criação da tabela de usuários
            cursor.execute(''' 
                CREATE TABLE IF NOT EXISTS usuario (
                    usuario_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username VARCHAR(100) NOT NULL UNIQUE,
                    password VARCHAR(50) NOT NULL
                ) 
            ''')
            print("Tabela 'usuario' criada com sucesso.")

            # Criação da tabela de categorias
            cursor.execute(''' 
                CREATE TABLE IF NOT EXISTS categoria (
                    categoria_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome_categoria VARCHAR(50) NOT NULL UNIQUE,
                    descricao_categoria TEXT,
                    ativo BOOLEAN DEFAULT 1
                ) 
            ''')
            print("Tabela 'categoria' criada com sucesso.")

            # Popula a tabela de categorias com valores iniciais
            cursor.executemany('''
                INSERT OR IGNORE INTO categoria (categoria_id, nome_categoria, descricao_categoria) 
                VALUES (?, ?, ?)
            ''', [
                (1, 'Livraria', 'Produtos relacionados a livros, material escolar e papelaria'),
                (2, 'Roupas', 'Vestuário em geral incluindo camisas, calças e vestidos'),
                (3, 'Acessórios', 'Complementos como bolsas, cintos e bijuterias')
            ])
            print("Categorias inseridas com sucesso.")

            # Criação da tabela de produtos
            cursor.execute(''' 
                CREATE TABLE IF NOT EXISTS produtos (
                    produto_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome_produto TEXT NOT NULL,
                    descricao_produto TEXT,
                    preco REAL NOT NULL CHECK(preco >= 0),
                    quantidade_estoque INTEGER NOT NULL CHECK(quantidade_estoque >= 0),
                    categoria_id INTEGER NOT NULL,
                    data_cadastro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (categoria_id) REFERENCES categoria(categoria_id)
                ) 
            ''')
            print("Tabela 'produtos' criada com sucesso.")

            # Criação da tabela de métodos de pagamento
            cursor.execute(''' 
                CREATE TABLE IF NOT EXISTS tipo_pagamento (
                    pagamento_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    metodo_pagamento TEXT NOT NULL UNIQUE
                ) 
            ''')
            print("Tabela 'tipo_pagamento' criada com sucesso.")


            # Inserção de métodos de pagamento
            cursor.executemany('''
                INSERT OR IGNORE INTO tipo_pagamento (pagamento_id, metodo_pagamento) 
                VALUES (?, ?)
            ''', [(0, 'Dinheiro'), (1, 'Multibanco'), (2, 'MB Way')])
            print("Métodos de pagamento inseridos com sucesso.")
            

            # Criação da tabela de vendas
            cursor.execute(''' 
                CREATE TABLE IF NOT EXISTS vendas (
                    venda_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    total REAL NOT NULL CHECK(total >= 0),
                    pagamento_id INTEGER NOT NULL,
                    data_venda TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (pagamento_id) REFERENCES tipo_pagamento(pagamento_id)
                ) 
            ''')
            print("Tabela 'vendas' criada com sucesso.")

            # Criação da tabela de itens de venda
            cursor.execute(''' 
                    CREATE TABLE IF NOT EXISTS itens_venda (
                    item_venda_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    venda_id INTEGER NOT NULL,
                    produto_id INTEGER NOT NULL,
                    quantidade INTEGER NOT NULL CHECK(quantidade > 0),
                    preco_unitario REAL NOT NULL CHECK(preco_unitario >= 0),
                    FOREIGN KEY (venda_id) REFERENCES vendas(venda_id),
                    FOREIGN KEY (produto_id) REFERENCES produtos(produto_id)
                ) 
            ''')
            print("Tabela 'itens_venda' criada com sucesso.")




            # Criação da tabela de inventário   
            cursor.execute('''  
                    CREATE TABLE IF NOT EXISTS  inventario (
                           inventario_id INTEGER PRIMARY KEY AUTOINCREMENT,
                           produto_id INTEGER NOT NULL,
                           quantidade INTEGER NOT NULL,
                           tipo_movimentacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                           data_movimentacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                           observacao TEXT, 
                           FOREIGN KEY (produto_id) REFERENCES produtos(produto_id)
                    )
         ''')
            print("Tabela 'inventario' criada com sucesso.")


            conn.commit()
            print("Base de dados criada com sucesso!")

    except sqlite3.Error as e:
        print(f"Erro ao criar banco de dados: {e}")

        



# Cabeçalho
def cabecalho():
    titulo = "AGAPE SHOP"
    tam = int(len(titulo)/2)+2
    print('--' * tam)
    print(f'  {titulo}  ')
    print('--' * tam)
    sleep(0.5)


# Cadastrar Usuario
def cadastro():
    while True:
        global user_id
        print(f"Deseja fazer Login ou Cadastro?")
        print("1. Login")
        print("2. Cadastro")
        print("3. Sair ")
        

        try:
            opcao = int(input("Escolha uma opção: "))

            if opcao == 1:
                print(" --- LOGIN --- ")
                nome = input("\nDigite o seu nome : ")
                #senha = input("Digite a sua password: ") sem o getpass
                senha = pwinput.pwinput("Digite a sua senha: ", mask="*") 
                print("\nA validar...\n")
                sleep(1)
                
                #conexao ao db e verificacao de user 
                with sqlite3.connect('agapeshop.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT usuario_id FROM usuario WHERE nome = ? AND senha = ?", (nome,senha))
                    user = cursor.fetchone()
                    

                    if user:  # Verifica se o usuário existe
                        user_id = user[0]
                        print("Login efetuado com sucesso!\n")
                        sleep(0.5)
                        break
                    else:
                        print("Nome de usuário ou senha incorretos. Você precisa se cadastrar primeiro.")
                        print("Você deseja se cadastrar? (s/n)")
                        if input().strip().lower() == 's':
                            continue

                
            elif opcao == 2:
                print(" --- CADASTRO --- ")  
                nome = input("Digite seu nome de usuario: ").strip().title()
                senha = input("Digite sua senha: ")

                with sqlite3.connect('agapeshop.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO usuario (nome , senha) VALUES(?, ?)",(nome, senha))
                    conn.commit()
                    print(f"Usuario {nome} cadastrado com suceso! ")

                    return nome
                

            elif opcao == 3:
                print("Saindo... Até a próxima!")
                break

            else:
                print("Opção Inválida! Tente novamente.")

        

        except ValueError:
            print("Erro: Por favor, insira um número válido.")
    


# Insere uma categoria ** e para selecionar a categoria e nao inserir pois ja deifini as categorias na criacao 
def insertCategoria(categoria):
    try:
        with sqlite3.connect('agapeshop.db') as conn:
            cursor = conn.cursor()
            
            # Busca direta pelo ID da categoria
            cursor.execute("SELECT categoria_id FROM categoria WHERE livraria = ? OR roupas = ? OR acessorios = ?", 
                         (categoria, categoria, categoria))
            
            resultado = cursor.fetchone()
            if resultado:
                return resultado[0]
            
            print(f"Categoria {categoria} encontrada com sucesso!")
            return resultado[0]

    except sqlite3.Error as e:
        print(f"Erro ao verificar Categoria: {e}")   
          
        

# Função para inserir um produto
def inserirProduto(nome_produto, descricao_produto, preco, quantidade_estoque, categoria_id):
    try:
        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO produtos (nome_produto, descricao_produto, preco, quantidade_estoque, categoria_id)
                VALUES (?, ?, ?, ?, ?)
            ''', (nome_produto, descricao_produto, preco, quantidade_estoque, categoria_id))
            conn.commit()
    except sqlite3.Error as e:
        print(f"Erro ao inserir produto: {e}")
        raise
        


# Funcao para atualizar um produto
def update(produto_id, novo_nome, novo_preco, nova_quantidade, janela):
    try:
        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()

            # Validar se o produto existe
            cursor.execute("SELECT * FROM produtos WHERE produto_id = ?", (produto_id,))
            produto = cursor.fetchone()

            if not produto:
                messagebox.showerror("Erro", f"Produto com ID {produto_id} não encontrado.")
                return

            # Calcular a diferença na quantidade de estoque
            estoque_atual = produto[4]  # Quantidade em estoque na tabela 'produtos'
            diferenca = int(nova_quantidade) - estoque_atual

            # Atualizar os dados do produto
            cursor.execute(
                '''
                UPDATE produtos
                SET nome_produto = ?, preco = ?, quantidade_estoque = ?
                WHERE produto_id = ?
                ''',
                (novo_nome, novo_preco, nova_quantidade, produto_id),
            )

            # Registrar movimentação no inventário, se houver alteração na quantidade
            if diferenca != 0:
                tipo_movimentacao = "entrada" if diferenca > 0 else "saida"
                registrar_movimentacao(
                    produto_id=produto_id,
                    quantidade=abs(diferenca),
                    tipo_movimentacao=tipo_movimentacao,
                    observacao="Alteração manual no estoque"
                )

            conn.commit()
            messagebox.showinfo("Sucesso", f"Produto com ID {produto_id} atualizado com sucesso!")
            janela.destroy()

    except sqlite3.Error as e:
        messagebox.showerror("Erro", f"Erro ao atualizar produto: {e}")



#Mostar produtos 
def mostrarProdutos():
    try:
        with sqlite3.connect('agapeshop.db') as conn:
            cursor = conn.cursor()

            cursor.execute('SELECT produto_id, nome_produto, descricao_produto, preco, quantidade_estoque, categoria_id FROM produtos')
            produtos = cursor.fetchall()

            if not produtos:
                print("Nenhum produto encontrado .")
                return []
            
            #return produtos
            
        for produto in produtos:
            print(f"\nNome: {produto[1]} \nDescricão: {produto[2]} \nPreco: {produto[3]} €  \nEstoque: {produto[4]}\n ")  #0Categoria 1Nome do produto,3preco do produto, 4 estoque do produto
        return produtos

    except sqlite3.Error as e:
        print(f"Erro ao mostrar produtos {e}")        


#funcao para deletar um produto
def delete():
    try:
        nome_produto = input("Digite o nome do produto que deseja deletar: ").strip().title()
       
        with sqlite3.connect('agapeshop.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT produto_id, nome_produto, descricao_produto  FROM produtos WHERE nome_produto = ?", (nome_produto,))
            produtos = cursor.fetchall()

            if produtos:
                print("Produtos encontrados:")
                for produto in produtos:
                    print(f"ID: {produto[0]}, Nome: {produto[1]}, Descrição: {produto[2]}")
                
                # Solicita que o usuário escolha o produto pelo ID
                try:
                    produto_id = int(input("Digite o ID do produto que deseja deletar: "))
                    cursor.execute("SELECT * FROM produtos WHERE produto_id = ?", (produto_id,))
                    produto_selecionado = cursor.fetchone()

                    if produto_selecionado:
                        cursor.execute("DELETE FROM produtos WHERE produto_id = ?", (produto_id,))
                        conn.commit()
                        print(f"Produto com ID {produto_id} excluído com sucesso.")
                    else:
                        print(f"Nenhum produto encontrado com o ID {produto_id}.")
                
                except ValueError:
                    print("ID inválido. Por favor, insira um número inteiro.")

            else:
                print(f"Nenhum produto encontrado com o nome {nome_produto}.")

    except sqlite3.Error as e:
        print(f"Erro ao deletar produto: {e}")

                    
# Função para registrar um novo usuário
def registrar_usuario(username, password):
    """
    Registra um novo usuário no banco de dados.
    """
    try:
        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO usuario (username, password) VALUES (?, ?)", (username, password))
            conn.commit()
            return True  # Sucesso
    except sqlite3.IntegrityError:
        return False  # Usuário já existe    
    
    
# Função para validar login
def validar_usuario(username, password):
    """
    Valida as credenciais do usuário.
    """
    with sqlite3.connect("agapeshop.db") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM usuario WHERE username = ? AND password = ?", (username, password))
        return cursor.fetchone() is not None


# Função para buscar produtos por categoria NAO APLICADA ATE O MOMENTO
def buscarPorCategoria(categoria_id):
    """
    Retorna produtos de uma categoria específica.
    """
    with sqlite3.connect("agapeshop.db") as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT nome_produto, descricao_produto, preco, quantidade_estoque 
            FROM produtos WHERE categoria_id = ?
        """, (categoria_id,))
        return cursor.fetchall()    
                    

def exportar_relatorio_pdf(relatorio_funcao, titulo="Relatório de Vendas"):
    relatorio = relatorio_funcao()

    if not relatorio:
        messagebox.showerror("Erro", "Não há dados para exportar.")
        return
    
        # Escolher o local para salvar o PDF
    arquivo = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")],
        title="Salvar Relatório como"
    )

    if not arquivo:  # Se o usuário cancelar
        return
    
    #incialização do pdf
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    #Titulo do relatório
    pdf.set_font("Arial", size=16, style="B")
    pdf.cell(200, 10, txt=titulo, ln=True, align="C")

    #Espaço
    pdf.ln(10)

    # Cabeçalhos do Relatório
    pdf.set_font("Arial", style="B", size=12)
    pdf.cell(30, 10, "ID Venda", border=1)
    pdf.cell(40, 10, "Produto", border=1)
    pdf.cell(30, 10, "Quantidade", border=1)
    pdf.cell(30, 10, "Preço Unitário", border=1)
    pdf.cell(30, 10, "Total", border=1)
    pdf.cell(30, 10, "Pagamento", border=1)
    pdf.ln()

    # Dados do Relatório
    pdf.set_font("Arial", size=12)
    for venda in relatorio:
        venda_id, total, metodo_pagamento, data_venda, nome_produto, quantidade, preco_unitario = venda
        pdf.cell(30, 10, str(venda_id), border=1)
        pdf.cell(40, 10, nome_produto, border=1)
        pdf.cell(30, 10, str(quantidade), border=1)
        pdf.cell(30, 10, f"€{preco_unitario:.2f}", border=1)
        pdf.cell(30, 10, f"€{quantidade * preco_unitario:.2f}", border=1)
        pdf.cell(30, 10, metodo_pagamento, border=1)
        pdf.ln()

        #Total da Geral
        pdf.ln(10)
        total_geral = sum(venda[5] * venda[6] for venda in relatorio)
        pdf.cell(200, 10, txt=f"Total Geral: €{total_geral:.2f}", ln=True, align="R")

        # salva o arquivo PDF
        pdf.output("relatorio_vendas.pdf")
        messagebox.showinfo("Sucesso", "Relatório exportado com sucesso!")

        #exportar o relatório
        exportar_relatorio_pdf(relatorio_diario, "Relatório de Vendas")



def relatorio_diario():
    """
    Retorna um relatório com todas as vendas realizadas no dia atual,
    incluindo tipo de pagamento.
    """
    try:
        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()
            # Consulta ajustada para incluir o tipo de pagamento
            query = '''
                SELECT v.venda_id, v.total, tp.metodo_pagamento, v.data_venda,
                       p.nome_produto, iv.quantidade, iv.preco_unitario
                FROM vendas v
                JOIN itens_venda iv ON v.venda_id = iv.venda_id
                JOIN produtos p ON iv.produto_id = p.produto_id
                JOIN tipo_pagamento tp ON v.pagamento_id = tp.pagamento_id
                WHERE DATE(v.data_venda) = DATE('now')
                ORDER BY v.data_venda DESC
            '''
            cursor.execute(query)
            return cursor.fetchall()
    except sqlite3.Error as e:
        print(f"Erro ao gerar relatório diário: {e}")
        return []


def consultar_inventario():
    try:
        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()
            cursor.execute(''' 
                                SELECT inventario_id, nome_produto, quantidade, tipo_movimentacao, data_movimentacao, observacao 
                                FROM inventario 
                                JOIN produtos p ON i.produto_id = p.produto_id
                                ORDER BY data_movimentacao DESC ''')
            
            resultados = cursor.fetchall()

            if not resultados:
                messagebox.showinfo("Inventário", "Nenhuma movimentação no inventário.")
                
                return
            
            return resultados

    except sqlite3.Error as e:
        messagebox.showerror(f"Erro ao consultar inventário:{e}")
                            


def exportar_inventario_excel():
    try:
        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()

            # Consulta corrigida para obter o inventário
            cursor.execute('''
                SELECT 
                    p.produto_id, 
                    p.nome_produto, 
                    p.descricao_produto, 
                    p.preco, 
                    COALESCE(SUM(CASE WHEN i.tipo_movimentacao = 'entrada' THEN i.quantidade 
                                      WHEN i.tipo_movimentacao = 'saida' THEN -i.quantidade END), 0) AS estoque_atual
                FROM produtos p
                LEFT JOIN inventario i ON p.produto_id = i.produto_id
                GROUP BY p.produto_id, p.nome_produto, p.descricao_produto, p.preco;
            ''')

            inventario = cursor.fetchall()

        # Verificar se há dados
        if not inventario:
            messagebox.showinfo("Exportar Inventário", "Nenhum dado encontrado no inventário.")
            return

        # Escolher o arquivo para salvar
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar Inventário"
        )
        if not arquivo:
            return

        # Criar o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventário Atual"

        # Cabeçalhos
        sheet.append(["ID Produto", "Nome", "Descrição", "Preço (€)", "Estoque Atual"])

        # Adicionar dados ao Excel
        for row in inventario:
            sheet.append(row)

        workbook.save(arquivo)
        messagebox.showinfo("Sucesso", f"Inventário exportado com sucesso para:\n{arquivo}")

    except sqlite3.Error as e:
        messagebox.showerror("Erro", f"Erro ao consultar o inventário: {e}")
    except PermissionError:
        messagebox.showerror("Erro", "Permissão negada. Feche o arquivo se ele estiver aberto e tente novamente.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exportar inventário: {e}")



def consultar_estoque_por_data(data_limite):
    try:
        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()

            cursor.execute('''
                SELECT 
                    p.produto_id, 
                    p.nome_produto, 
                    p.descricao_produto, 
                    p.preco, 
                    p.quantidade_estoque 
                    + COALESCE(SUM(CASE 
                        WHEN i.tipo_movimentacao = 'entrada' THEN i.quantidade
                        WHEN i.tipo_movimentacao = 'saida' THEN -i.quantidade
                        ELSE 0 END), 0) AS estoque_atual
                FROM produtos p
                LEFT JOIN inventario i 
                    ON p.produto_id = i.produto_id 
                    AND i.data_movimentacao <= ?
                GROUP BY p.produto_id, p.nome_produto, p.descricao_produto, p.preco, p.quantidade_estoque
                ORDER BY p.produto_id;
            ''', (data_limite,))
            return cursor.fetchall()
    except sqlite3.Error as e:
        print(f"Erro ao consultar estoque por data: {e}")
        return []



def registrar_movimentacao(produto_id, quantidade, tipo_movimentacao, observacao=""):
    """
    Registra uma movimentação no inventário.
    """
    try:
        if tipo_movimentacao not in ["entrada", "saida"]:
            raise ValueError("O tipo de movimentação deve ser 'entrada' ou 'saida'.")

        if quantidade <= 0:
            raise ValueError("A quantidade deve ser maior que 0.")

        with sqlite3.connect("agapeshop.db") as conn:
            cursor = conn.cursor()

            # SQL correto para registrar a movimentação
            cursor.execute('''
                INSERT INTO inventario (produto_id, quantidade, tipo_movimentacao, data_movimentacao, observacao)
                VALUES (?, ?, ?, datetime('now'), ?)
            ''', (produto_id, quantidade, tipo_movimentacao, observacao))
            
            conn.commit()
            print("Movimentação registrada no inventário com sucesso.")
    except sqlite3.Error as e:
        print(f"Erro ao registrar movimentação no inventário: {e}")


